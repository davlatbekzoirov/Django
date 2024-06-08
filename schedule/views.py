from django.shortcuts import render
from .forms import ScheduleForm
import openpyxl
from django.http import JsonResponse
import datetime
from .models import Holiday, LabTemperature

def dashboard(request):
    labs = LabTemperature.objects.values_list('lab_id', flat=True).distinct()
    return render(request, 'dashboard.html', {'labs': labs})

def get_lab_data(request, lab_id):
    lab_data = LabTemperature.objects.filter(lab_id=lab_id).order_by('date', 'time')
    data = {
        'dates': [entry.date.strftime('%Y-%m-%d') for entry in lab_data],
        'times': [entry.time.strftime('%H:%M:%S') for entry in lab_data],
        'temp_inside': [entry.temp_inside for entry in lab_data],
        'temp_outside': [entry.temp_outside for entry in lab_data],
        'temp_dif': [entry.temp_dif for entry in lab_data],
    }
    return JsonResponse(data)


def schedule_view(request):
    """
    Dars jadvalini ko'rish va yuklash uchun ko'rsatma.
    """
    if request.method == 'POST':
        form = ScheduleForm(request.POST, request.FILES)
        if form.is_valid():
            # Foydalanuvchi ID sini so'roqdan yoki sessiyadan olish
            user_id = request.user.id  #  Foydalanuvchi tasdiqlangan bo'lsa

            # Joriy foydalanuvchi uchun eski yozuvlarni o'chirish
            Holiday.objects.filter(user_id=user_id).delete()

            # Yuklangan faylni ishlash
            schedule_file = request.FILES['file']
            sheets_data = parse_schedule(schedule_file)
            sheet_name = request.POST.get('sheet_name', list(sheets_data.keys())[0])
            schedule_data = sheets_data[sheet_name]

            holidays = form.cleaned_data['holidays']

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'schedule_data': schedule_data, 'holidays': holidays})

            context = {
                'form': form,
                'schedule_file': schedule_file,
                'sheets_data': sheets_data,
                'schedule_data': schedule_data,
                'from_month': dict(form.fields['from_month'].choices)[form.cleaned_data['from_month']],
                'to_month': dict(form.fields['to_month'].choices)[form.cleaned_data['to_month']],
                'holidays': holidays,
            }
            return render(request, 'index.html', context)
    else:
        form = ScheduleForm()
    return render(request, 'index.html', {'form': form})

def parse_schedule(file):
    """
    Excel faylni o'qish va jadval ma'lumotlarini ajratish.
    """
    wb = openpyxl.load_workbook(file)
    sheets_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        schedule = []
        for row in ws.iter_rows(values_only=True):
            formatted_row = []
            for cell in row:
                if isinstance(cell, datetime.time):
                    formatted_row.append(cell.strftime('%H:%M'))
                elif cell is None:
                    formatted_row.append('')
                else:
                    formatted_row.append(cell)
            schedule.append(formatted_row)
        sheets_data[sheet_name] = schedule

    return sheets_data
